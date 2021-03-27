from io import StringIO ## for Python 3
import io
import openpyxl
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl.styles.borders import Border, Side
from openpyxl.styles import  Font
import os


class Export:
	def __init__(self, data, ruta):
		self.data = data
		self.ruta = ruta

	def plantilla(self):
		print(openpyxl.__version__)

	def toExcel(self):
		print(self.ruta)
		# print(self.ruta, 'desde toExcel.py')
		self.wb = Workbook()
		self.ws = self.wb.active # worksheet
		########## Insertamos datos ###############

		self.my_png = openpyxl.drawing.image.Image(self.ruta)
		self.ws.add_image(self.my_png, 'A1')

		self.titulo = Font(
			name='Calibri',
			size=16,
			bold=True,
			italic=False,
			vertAlign=None,
			underline='none',
			strike=False,
			color='000000')


		self.border = Border(left=Side(style='thin'), 
            right=Side(style='thin'), 
            top=Side(style='thin'), 
            bottom=Side(style='thin'))

		self.ws['A6'].value = 'Usted Litiga, Nosotros Liquidamos'
		self.ws['A6'].font = self.titulo

		self.ws['D2'].border = self.border
		self.ws['D2'].value = 'Hugo lo estamos logrando'

		######################################################### 

		############# Aqui diseniamos el archivo  #################
		self.ws.title = "Resultado caso-" + self.data[2]['caso']
		############ aqui se importa a el frontend ################
		# self.wb.save("test.xlsx")
		self.out = io.BytesIO()
		self.wb.save(self.out)
		self.out.seek(0)
		return self.out 
